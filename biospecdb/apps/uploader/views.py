from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
from django.db import transaction
from openpyxl import load_workbook

from .forms import FileUploadForm, DataInputForm
from uploader.models import Patient, Visit, SpectralData, BioSample, Symptom, Disease
from biospecdb.util import is_valid_uuid, to_uuid

def home(request):
    return render(request, 'home.html')


@staff_member_required
def upload_file(request):
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            #return render(request, 'UploadSuccess.html')
            return display_xlsx(request)
    else:
        form = FileUploadForm()
    return render(request, 'MetadataFileUpload.html', {'form': form})


@staff_member_required
def display_xlsx(request):
    workbook = load_workbook('./biospecdb/apps/uploader/uploads/METADATA_barauna2021ultrarapid.xlsx')
    worksheet = workbook.active
    data = []
    for row in worksheet.iter_rows(values_only=True):
        data.append(row)
    return render(request, 'MetadataDisplay.html', {'data': data})


@staff_member_required
def data_input(request):
    message = ""
    form = DataInputForm()
    delta_count = len(form.base_fields) - 1
    
    if request.method == 'POST':
        form = DataInputForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()  # Save data to database.
            patient_id = form.cleaned_data["patient_id"]
            message = "Data Input with Patient ID {} has been submitted successfully!!!".format(patient_id)
            return render(request, 'DataInputForm.html', {'form': form, 'message': message, 'delta_count': delta_count})
        
    elif request.method == 'GET':
        form = DataInputForm()
        patient_id = request.GET.get('patient_id')
        if patient_id:
            if not is_valid_uuid(patient_id):
                message = "The provided Patient ID {} is not a valid number.".format(patient_id)
                return render(request, 'DataInputForm.html', {'form': form, 'message': message, \
                    'delta_count': delta_count})
            else:
                patient_id = to_uuid(patient_id)
                with transaction.atomic():
                    try:
                        patient = Patient.objects.select_for_update().get(patient_id=patient_id)
                    except (Patient.DoesNotExist):
                        message = "Data Search failed - there is no data associated with Patient ID {}." \
                            .format(patient_id)
                        return render(request, 'DataInputForm.html', {'form': form, 'message': message, \
                            'delta_count': delta_count})
                    try:
                        last_visit = Visit.objects.select_for_update().filter(patient_id=patient_id) \
                            .order_by('created_at').last()
                    except (Visit.DoesNotExist):
                        message = "Data Search failed - there is no any visit of patient with Patient ID {}." \
                            .format(patient_id)
                        return render(request, 'DataInputForm.html', {'form': form, 'message': message, \
                            'delta_count': delta_count})
                    try:
                        biosample = BioSample.objects.select_for_update().get(visit=last_visit)
                    except (BioSample.DoesNotExist):
                        message = "Data Search failed - there is no biosample associated with the visit {}." \
                            .format(last_visit)
                        return render(request, 'DataInputForm.html', {'form': form, 'message': message, \
                            'delta_count': delta_count})
                    try:
                        last_visit_symptoms = Symptom.objects.select_for_update().filter(visit=last_visit)
                        symptom = last_visit_symptoms.order_by('days_symptomatic').last()
                    except (Symptom.DoesNotExist):
                        message = "Data Search failed - there are no symptoms associated with the visit {}." \
                            .format(last_visit)
                        return render(request, 'DataInputForm.html', {'form': form, 'message': message, \
                            'delta_count': delta_count})
                    try:
                        spectraldata = SpectralData.objects.select_for_update().get(bio_sample=biosample)
                    except (SpectralData.DoesNotExist):
                        message = "Data Search failed - there is no spectral data associated with the biosample {}." \
                            .format(biosample)
                        return render(request, 'DataInputForm.html', {'form': form, 'message': message, \
                            'delta_count': delta_count})
                
                    initial_data={
                        'patient_id': patient_id,
                        'gender': patient.gender,
                        'days_symptomatic': symptom.days_symptomatic,
                        'patient_age': last_visit.patient_age,
                        'spectra_measurement': spectraldata.spectra_measurement,
                        'instrument': spectraldata.instrument,
                        'acquisition_time': spectraldata.acquisition_time,
                        'n_coadditions': spectraldata.n_coadditions,
                        'resolution': spectraldata.resolution,
                        'sample_type': biosample.sample_type,
                        'sample_processing': biosample.sample_processing,
                        'freezing_temp': biosample.freezing_temp,
                        'thawing_time': biosample.thawing_time,
                        'spectral_data': spectraldata.data
                    }
                    for symptom in last_visit_symptoms:
                        if symptom.disease.value_class == "BOOL":
                            initial_data[symptom.disease.name] = \
                                Disease.Types(symptom.disease.value_class).cast(symptom.disease_value)   
                        else:
                            initial_data[symptom.disease.name] = symptom.disease_value
                    form = DataInputForm(initial=initial_data)
                    message = "The data associated with Patient ID {} is shown below:".format(patient_id)
                    return render(request, 'DataInputForm.html', {'form': form, 'message': message, \
                        'delta_count': delta_count})
                
    else:
        form = DataInputForm()
        
    return render(request, 'DataInputForm.html', {'form': form, 'message': message, 'delta_count': delta_count})
